from datetime import datetime as dt
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import MetaTrader5 as mt5
import pandas as pd
import pytz
import os

pd.set_option('display.max_columns', 500) # number of columns to be displayed
pd.set_option('display.max_rows', 1500)   # max table width to display


stat_data = {}

def get_data(start_datetime, num_bars, symbol, time_frame):
    mt5.initialize()
    timezone = pytz.timezone('Etc/UTC')
    utc_from = dt(start_datetime.year, start_datetime.month, start_datetime.day, 00, 00, tzinfo=timezone)
    rates = mt5.copy_rates_from(symbol, time_frame, utc_from, num_bars)
    mt5.shutdown()

    return rates


def generate_dictonary_values(start_datetime, symbol):

    tflist = [mt5.TIMEFRAME_D1, mt5.TIMEFRAME_W1, mt5.TIMEFRAME_MN1] #Time frames list
    nbarslist = [1673, 418, 55] #get 4.5 year data for calculations
    std_values = []

    j = 0
    while j < 3:

        rates = get_data(start_datetime,  nbarslist[j], symbol, tflist[j])

        if rates is not None:
            df = pd.DataFrame(rates)
            df['time'] = pd.to_datetime(df['time'], unit='s') 
            df = df.drop_duplicates(keep=False)
            df = df[['time', 'open', 'high', 'low', 'close']]
            df['open_to_open'] = df.open.pct_change() * 100    #precentage change 
            df['dif'] = df.open.diff()                         #change in number

            # print(df)
            statistics = df.describe().open_to_open 
            #print(statistics)

            #These data will embered to dictonary
            std = statistics.tolist()[2]
            std_values.append(std)

        j += 1

    #adding data to dictonary
    stat_data[symbol] = std_values



def push_data_to_exel():

    #sort all tradable currencies accoding to standered devition value
    sorted_all_tradable_currencies = {k: v for k, v in sorted(stat_data.items(), key=lambda item: item[1], reverse=True)} 

    currentpath = r'C:\Users\user\Desktop\Global Macro\G10\Volatility Analysis'
    filename = 'std_Analysis.xlsx'
    fullpath = os.path.join(currentpath,filename)
    #print(full_path)

    #open exel workbook
    wb = openpyxl.load_workbook(fullpath)
    #print(wb.sheetnames)

    #Access into worksheet
    sheet = wb['std_Summary']

    #insert data to exel sheet
    for row,i in enumerate(sorted_all_tradable_currencies):
        sheet.cell(row=row+2, column=1).value = i  #currency name
        sheet.cell(row=row+2, column=2).value = sorted_all_tradable_currencies[i][0] #standerd deviation daily value
        sheet.cell(row=row+2, column=3).value = sorted_all_tradable_currencies[i][1] #standerd deviation weekly value
        sheet.cell(row=row+2, column=4).value = sorted_all_tradable_currencies[i][2] #standerd deviation monthly value

        #decide colour accouding to urrency is major or not
        if currency_is_major(i): 
            sheet.cell(row=row+2, column=1).fill = PatternFill("solid", fgColor="0099CCFF")

        else:
            sheet.cell(row=row+2, column=1).fill = PatternFill("solid", fgColor="00FFCC99")

    wb.save(fullpath)



def  currency_is_major(currency_pair):
    tset = ['AUDCAD', 'AUDCHF', 'AUDJPY', 'AUDNZD', 'AUDUSD', 'CADCHF', 'CADJPY', 'CHFJPY', 'EURAUD', 'EURCAD', 'EURCHF', 'EURGBP', 
            'EURJPY''EURNZD', 'EURUSD', 'GBPAUD', 'GBPCAD', 'GBPCHF', 'GBPJPY',  'GBPNZD', 'GBPUSD', 'NZDCAD', 'NZDCHF', 'NZDJPY', 
            'NZDUSD', 'USDCAD', 'USDCHF', 'USDJPY', 'EURNOK', 'EURSEK','GBPNOK', 'GBPSEK', 'NOKJPY', 'NOKSEK', 'SEKJPY', 'USDNOK', 'USDSEK']

    for i in tset:
        if currency_pair == i:
            condition = True
            break
        else:
            condition = False

    return condition
    


if __name__ == '__main__':
    
    start_datetime = dt(2020, 4, 1) 

    tradable_set = ['AUDCAD', 'AUDCHF', 'AUDJPY', 'AUDNZD', 'AUDUSD', 'CADCHF', 'CADJPY', 'CHFJPY', 'EURAUD', 'EURCAD', 
                    'EURCHF', 'EURGBP', 'EURJPY', 'EURNOK', 'EURNZD', 'EURPLN', 'EURSEK', 'EURUSD', 'EURZAR', 'GBPAUD', 
                    'GBPCAD', 'GBPCHF', 'GBPJPY', 'GBPNOK', 'GBPNZD', 'GBPSEK', 'GBPTRY', 'GBPUSD', 'NOKJPY', 'NOKSEK', 
                    'NZDCAD', 'NZDCHF', 'NZDJPY', 'NZDUSD', 'SEKJPY', 'USDCAD', 'USDCHF', 'USDCZK', 'USDHUF', 'USDJPY', 
                    'USDMXN', 'USDNOK', 'USDPLN', 'USDRUB', 'USDSEK', 'USDTRY', 'USDZAR']

    i = 0
    while i < 47:

        generate_dictonary_values(start_datetime, tradable_set[i])
        print(i+1, tradable_set[i], ' is processing....')
        i += 1

    push_data_to_exel()











